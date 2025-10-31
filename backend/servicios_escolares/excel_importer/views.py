import os
import logging
from decimal import Decimal, ROUND_DOWN
from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import ForeignKey
from django.contrib.auth.decorators import login_required
from .forms import UploadExcelForm, MapeoCamposForm
from .models import ModeloAutorizado
from .utils import obtener_instancia_relacionada
import openpyxl
from openpyxl.utils import range_boundaries

logger = logging.getLogger(__name__)


def leer_rango_excel(ruta_archivo, nombre_hoja, rango_excel):
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    if nombre_hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{nombre_hoja}' no existe en el archivo.")
    ws = wb[nombre_hoja]

    min_col, min_row, max_col, max_row = range_boundaries(rango_excel)

    merged_cells_map = {}
    for merged_range in ws.merged_cells.ranges:
        minc, minr, maxc, maxr = merged_range.bounds
        valor = ws.cell(row=minr, column=minc).value
        for r in range(minr, maxr + 1):
            for c in range(minc, maxc + 1):
                merged_cells_map[(r, c)] = valor

    datos = []
    for row in range(min_row, max_row + 1):
        fila = []
        for col in range(min_col, max_col + 1):
            valor = merged_cells_map.get((row, col), ws.cell(row=row, column=col).value)
            fila.append(valor)
        datos.append(fila)

    return datos


def normalizar_choice(valor, field):
    if valor is None:
        return None
    valor_str = str(valor).strip().lower()
    for choice_val, _ in field.choices:
        if choice_val.lower() == valor_str:
            return choice_val
    return None

@login_required
def importar_modelo(request, pk):
    modelo_autorizado = get_object_or_404(ModeloAutorizado, pk=pk)
    ModelClass = modelo_autorizado.get_model_class()
    campos_modelo = [f.name for f in ModelClass._meta.get_fields() if f.concrete and not f.auto_created]
    contexto = {'modelo_autorizado': modelo_autorizado}

    try:
        if request.method == 'POST':
            if 'subir_excel' in request.POST:
                form = UploadExcelForm(request.POST, request.FILES)
                if form.is_valid():
                    archivo = request.FILES['archivo']
                    hoja = form.cleaned_data['hoja']
                    rango = form.cleaned_data['rango']
                    ruta = os.path.join(settings.MEDIA_ROOT, archivo.name)

                    with open(ruta, 'wb+') as destino:
                        for chunk in archivo.chunks():
                            destino.write(chunk)

                    datos_leidos = leer_rango_excel(ruta, hoja, rango)
                    encabezados = datos_leidos[0]
                    filas = datos_leidos[1:]

                    logger.info(f"Archivo leído correctamente: {len(filas)} filas encontradas.")
                    messages.info(request, f"Archivo leído correctamente: {len(filas)} filas encontradas.")

                    mapeo_form = MapeoCamposForm(campos_modelo, encabezados)

                    # Asegurar que lo guardado en sesión sea JSON-serializable
                    from datetime import datetime, date
                    from decimal import Decimal

                    def _json_safe(v):
                        if isinstance(v, (datetime, date)):
                            return v.isoformat()
                        if isinstance(v, Decimal):
                            try:
                                return float(v)
                            except Exception:
                                return str(v)
                        if isinstance(v, bytes):
                            try:
                                return v.decode('utf-8', errors='ignore')
                            except Exception:
                                return str(v)
                        return v

                    def _safe_row(row):
                        return [_json_safe(x) for x in row]

                    filas_seguras = [_safe_row(r) for r in filas]
                    encabezados_seg = [_json_safe(h) for h in encabezados]

                    request.session['datos_importacion'] = {
                        'ruta': ruta,
                        'hoja': hoja,
                        'rango': rango,
                        'encabezados': encabezados_seg,
                        'filas': filas_seguras,
                        'modelo_pk': pk,
                    }

                    contexto.update({
                        'mapeo_form': mapeo_form,
                        'encabezados': encabezados,
                        'filas_muestra': filas[:5],
                    })
                    return render(request, 'excel_importer/importar_generico.html', contexto)

                else:
                    messages.error(request, "Formulario inválido.")
                contexto['form'] = form

            elif 'importar_datos' in request.POST:
                datos_sesion = request.session.get('datos_importacion')
                if not datos_sesion:
                    messages.error(request, "No hay datos para importar.")
                    logger.error("No hay datos en sesión para importar.")
                    return redirect('importar_modelo', pk=pk)

                mapeo_form = MapeoCamposForm(campos_modelo, datos_sesion['encabezados'], request.POST)
                if mapeo_form.is_valid():
                    asignacion = mapeo_form.cleaned_data
                    encabezados = datos_sesion['encabezados']
                    filas = datos_sesion['filas']
                    col_indices = {campo: encabezados.index(col) for campo, col in asignacion.items()}

                    registros_importados = 0
                    filas_invalidas = []
                    filas_omitidas_fk = 0

                    campos_unicos = [f.name for f in ModelClass._meta.get_fields() if getattr(f, 'unique', False)]

                    for idx, fila in enumerate(filas):
                        datos = {}
                        error_en_fila = False
                        razon_error = ""

                        logger.debug(f"Procesando fila {idx + 2}: {fila}")

                        # Validar campos obligatorios
                        for campo in col_indices:
                            valor = fila[col_indices[campo]]
                            field = ModelClass._meta.get_field(campo)
                            if (valor in (None, '')) and not (field.null or field.blank):
                                error_en_fila = True
                                razon_error = f"Campo obligatorio '{campo}' vacío."
                                break
                        if error_en_fila:
                            filas_invalidas.append((idx, fila, razon_error))
                            logger.warning(f"Fila {idx + 2} omitida: {razon_error}")
                            continue

                        # Procesar campos y relaciones
                        for campo, idx_col in col_indices.items():
                            valor_celda = fila[idx_col]
                            field = ModelClass._meta.get_field(campo)

                            if isinstance(field, ForeignKey):
                                instancia = obtener_instancia_relacionada(field.related_model, valor_celda)
                                if instancia is None:
                                    filas_omitidas_fk += 1
                                    error_en_fila = True
                                    razon_error = f"No se encontró instancia relacionada para FK '{campo}' con valor '{valor_celda}'."
                                    break
                                datos[campo] = instancia
                            else:
                                try:
                                    if valor_celda in (None, ''):
                                        datos[campo] = None
                                    elif field.choices:
                                        valor_normalizado = normalizar_choice(valor_celda, field)
                                        if valor_normalizado is None:
                                            error_en_fila = True
                                            razon_error = f"Valor '{valor_celda}' no válido para campo con choices '{campo}'."
                                            break
                                        datos[campo] = valor_normalizado
                                    elif field.get_internal_type() == 'DecimalField':
                                        d = Decimal(str(valor_celda))
                                        d = d.quantize(Decimal(f'1.{"0"*field.decimal_places}'), rounding=ROUND_DOWN)
                                        datos[campo] = d
                                    elif field.get_internal_type() == 'IntegerField':
                                        datos[campo] = int(float(valor_celda))
                                    else:
                                        datos[campo] = valor_celda
                                except Exception as e:
                                    error_en_fila = True
                                    razon_error = f"Error al convertir campo '{campo}': {e}"
                                    break

                        if error_en_fila:
                            filas_invalidas.append((idx, fila, razon_error))
                            logger.warning(f"Fila {idx + 2} omitida: {razon_error}")
                            continue

                        # Evitar duplicados - modificado para manejar materias con clave única
                        if ModelClass.__name__ == 'Materia':
                            # Para materias, solo verificar por clave única
                            filtro_unico = {'clave': datos.get('clave')} if 'clave' in datos else {}
                            if filtro_unico and ModelClass.objects.filter(**filtro_unico).exists():
                                # Si ya existe, obtener la materia existente y crear/actualizar la relación MateriaCarrera
                                from datos_academicos.models import MateriaCarrera
                                materia_existente = ModelClass.objects.get(**filtro_unico)
                                
                                if 'carreras' in datos and datos['carreras']:
                                    carrera = datos['carreras']
                                    semestre = datos.get('semestre')  # Obtener semestre del Excel
                                    
                                    # Crear o actualizar la relación MateriaCarrera
                                    materia_carrera, created = MateriaCarrera.objects.get_or_create(
                                        materia=materia_existente,
                                        carrera=carrera,
                                        defaults={'semestre': semestre}
                                    )
                                    
                                    if not created and semestre is not None:
                                        # Si ya existía pero el semestre es diferente, actualizarlo
                                        if materia_carrera.semestre != semestre:
                                            materia_carrera.semestre = semestre
                                            materia_carrera.save()
                                            logger.info(f"Semestre actualizado para {materia_existente.clave} en {carrera}: {semestre}")
                                    
                                    action = "creada" if created else "actualizada"
                                    logger.info(f"Relación {action} para materia {materia_existente.clave} con carrera {carrera}")
                                
                                filas_invalidas.append((idx, fila, f"Materia con clave {datos.get('clave')} ya existe - relación procesada"))
                                continue
                        else:
                            # Para otros modelos, usar la lógica original
                            filtro_unico = {campo: datos[campo] for campo in campos_unicos if campo in datos}
                            if filtro_unico and ModelClass.objects.filter(**filtro_unico).exists():
                                filas_invalidas.append((idx, fila, f"Registro duplicado con {filtro_unico}"))
                                logger.info(f"Fila {idx + 2} omitida: registro duplicado con {filtro_unico}")
                                continue

                        try:
                            obj = ModelClass(**datos)
                            obj.full_clean()
                            obj.save()
                            
                            # Para materias, crear la relación MateriaCarrera después de guardar
                            if ModelClass.__name__ == 'Materia' and 'carreras' in datos and datos['carreras']:
                                from datos_academicos.models import MateriaCarrera
                                carrera = datos['carreras']
                                semestre = datos.get('semestre')
                                
                                MateriaCarrera.objects.create(
                                    materia=obj,
                                    carrera=carrera,
                                    semestre=semestre
                                )
                                logger.info(f"Relación creada para nueva materia {obj.clave} con carrera {carrera}, semestre {semestre}")
                            
                            registros_importados += 1
                        except ValidationError as e:
                            filas_invalidas.append((idx, fila, f"Error de validación: {e}"))
                            logger.warning(f"Fila {idx + 2} omitida por validación: {e}")
                        except Exception as e:
                            filas_invalidas.append((idx, fila, f"Error inesperado: {e}"))
                            logger.error(f"Fila {idx + 2} omitida por excepción inesperada", exc_info=True)

                    for idx, fila, error in filas_invalidas:
                        messages.warning(request, f"Fila {idx + 2} omitida: {error}")

                    if registros_importados > 0:
                        messages.success(request, f"Importados {registros_importados} registros.")
                    else:
                        if filas_invalidas:
                            messages.warning(request, f"No se importó ningún registro válido. {len(filas_invalidas)} filas fueron omitidas por errores.")
                        else:
                            messages.warning(request, "No se importó ningún registro. Verifica el archivo y el mapeo.")

                    if filas_omitidas_fk > 0:
                        messages.warning(request, f"Omitidas {filas_omitidas_fk} filas por claves foráneas inválidas.")

                    request.session.pop('datos_importacion', None)
                    return redirect('importar_modelo', pk=pk)

                messages.error(request, "Formulario de mapeo inválido.")
                encabezados = datos_sesion['encabezados']
                filas = datos_sesion['filas']
                contexto.update({
                    'mapeo_form': mapeo_form,
                    'encabezados': encabezados,
                    'filas_muestra': filas[:5]
                })

        else:
            contexto['form'] = UploadExcelForm()

    except Exception as e:
        messages.error(request, f"Se produjo un error inesperado: {str(e)}")
        logger.error(f"Error en importar_modelo: {e}", exc_info=True)

    return render(request, 'excel_importer/importar_generico.html', contexto)

@login_required
def importar_excel(request):
    contexto = {}
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            hoja = form.cleaned_data['hoja']
            rango = form.cleaned_data['rango']
            ruta = os.path.join(settings.MEDIA_ROOT, archivo.name)

            with open(ruta, 'wb+') as destino:
                for chunk in archivo.chunks():
                    destino.write(chunk)

            try:
                datos = leer_rango_excel(ruta, hoja, rango)
                contexto.update({
                    'datos': datos,
                    'rango': rango,
                    'hoja': hoja
                })
            except Exception as e:
                messages.error(request, f"Error al leer archivo: {str(e)}")
        else:
            messages.error(request, "Formulario inválido.")
    else:
        form = UploadExcelForm()

    contexto['form'] = contexto.get('form', form)
    return render(request, 'excel_importer/importar.html', contexto)
