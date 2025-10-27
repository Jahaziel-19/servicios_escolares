"""
Vistas administrativas para gestionar registros de aspirantes del formulario público
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
import csv
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from .models import SolicitudAdmision, PeriodoAdmision, FichaAdmision
from .email_utils import (
    enviar_notificacion_cambio_estado, 
    enviar_ficha_por_email,
    enviar_notificacion_estado_solicitud
)

# Verificación centralizada para usuarios administrativos (incluye Servicios Escolares)
def _is_admin_user(user):
    return (
        user.is_staff or user.is_superuser or 
        user.groups.filter(name__in=['ServiciosEscolares','Servicios Escolares']).exists()
    )


@login_required
def admin_dashboard_publico(request):
    """Dashboard administrativo para registros públicos"""
    if not _is_admin_user(request.user):
        messages.error(request, 'Acceso restringido: solo personal de Servicios Escolares.')
        return redirect('datos_academicos:servicios_login')
    # Estadísticas generales
    periodo_actual = PeriodoAdmision.objects.filter(activo=True).first()
    
    if periodo_actual:
        solicitudes_periodo = SolicitudAdmision.objects.filter(periodo=periodo_actual)
        
        # Estadísticas básicas
        total_solicitudes = solicitudes_periodo.count()
        solicitudes_hoy = solicitudes_periodo.filter(
            fecha_registro__date=timezone.now().date()
        ).count()
        
        # Estadísticas por estado
        stats_estado = solicitudes_periodo.values('estado').annotate(
            count=Count('id')
        ).order_by('estado')
        
        # Solicitudes recientes (últimas 10)
        solicitudes_recientes = solicitudes_periodo.order_by('-fecha_registro')[:10]
        
        # Estadísticas por carrera
        stats_carrera = solicitudes_periodo.exclude(
            respuestas_json__carrera_interes__isnull=True
        ).extra(
            select={'carrera': "JSON_EXTRACT(respuestas_json, '$.carrera_interes')"}
        ).values('carrera').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # Gráfico de registros por día (últimos 30 días)
        fecha_inicio = timezone.now().date() - timedelta(days=30)
        registros_por_dia = []
        for i in range(30):
            fecha = fecha_inicio + timedelta(days=i)
            count = solicitudes_periodo.filter(fecha_registro__date=fecha).count()
            registros_por_dia.append({
                'fecha': fecha.strftime('%Y-%m-%d'),
                'count': count
            })
        
        context = {
            'periodo_actual': periodo_actual,
            'total_solicitudes': total_solicitudes,
            'solicitudes_hoy': solicitudes_hoy,
            'stats_estado': stats_estado,
            'solicitudes_recientes': solicitudes_recientes,
            'stats_carrera': stats_carrera,
            'registros_por_dia': registros_por_dia,
        }
    else:
        context = {
            'periodo_actual': None,
            'mensaje': 'No hay período de admisión activo'
        }
    
    return render(request, 'admision/admin/dashboard.html', context)


@login_required
def admin_solicitudes_publico(request):
    """Lista de solicitudes del formulario público con filtros"""
    if not _is_admin_user(request.user):
        messages.error(request, 'Acceso restringido: solo personal de Servicios Escolares.')
        return redirect('datos_academicos:servicios_login')
    # Filtros
    periodo_id = request.GET.get('periodo')
    estado = request.GET.get('estado')
    busqueda = request.GET.get('q')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    carrera = request.GET.get('carrera')
    
    # Query base
    solicitudes = SolicitudAdmision.objects.all().order_by('-fecha_registro')
    
    # Aplicar filtros
    if periodo_id:
        solicitudes = solicitudes.filter(periodo_id=periodo_id)
    
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    
    if busqueda:
        solicitudes = solicitudes.filter(
            Q(folio__icontains=busqueda) |
            Q(curp__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(respuestas_json__nombre__icontains=busqueda) |
            Q(respuestas_json__apellido_paterno__icontains=busqueda) |
            Q(respuestas_json__apellido_materno__icontains=busqueda)
        )
    
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_registro__date__gte=fecha_desde_dt)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_registro__date__lte=fecha_hasta_dt)
        except ValueError:
            pass
    
    if carrera:
        solicitudes = solicitudes.extra(
            where=["JSON_EXTRACT(respuestas_json, '$.carrera_interes') = %s"],
            params=[carrera]
        )
    
    # Paginación
    paginator = Paginator(solicitudes, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para filtros
    periodos = PeriodoAdmision.objects.filter(activo=True).order_by('-fecha_inicio')
    estados_choices = SolicitudAdmision.ESTADOS
    
    # Carreras disponibles (extraer de JSON)
    carreras_disponibles = SolicitudAdmision.objects.exclude(
        respuestas_json__carrera_interes__isnull=True
    ).extra(
        select={'carrera': "JSON_EXTRACT(respuestas_json, '$.carrera_interes')"}
    ).values_list('carrera', flat=True).distinct()
    
    context = {
        'page_obj': page_obj,
        'periodos': periodos,
        'estados_choices': estados_choices,
        'carreras_disponibles': sorted(set(carreras_disponibles)),
        'filtros': {
            'periodo_id': periodo_id,
            'estado': estado,
            'busqueda': busqueda,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'carrera': carrera,
        }
    }
    
    return render(request, 'admision/admin/solicitudes_publico.html', context)


@login_required
def admin_ver_solicitud_publico(request, folio):
    """Ver detalles de una solicitud específica"""
    if not _is_admin_user(request.user):
        messages.error(request, 'Acceso restringido: solo personal de Servicios Escolares.')
        return redirect('datos_academicos:servicios_login')
    solicitud = get_object_or_404(SolicitudAdmision, folio=folio)
    
    # Obtener ficha si existe
    ficha = None
    try:
        ficha = FichaAdmision.objects.get(solicitud=solicitud)
    except FichaAdmision.DoesNotExist:
        pass
    
    context = {
        'solicitud': solicitud,
        'ficha': ficha,
        'estados_choices': SolicitudAdmision.ESTADOS,
    }
    
    return render(request, 'admision/admin/ver_solicitud_publico.html', context)


@login_required
@require_http_methods(["POST"])
def admin_cambiar_estado_solicitud(request, folio):
    """Cambiar estado de una solicitud"""
    if not _is_admin_user(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    solicitud = get_object_or_404(SolicitudAdmision, folio=folio)
    
    nuevo_estado = request.POST.get('estado')
    comentario = request.POST.get('comentario', '')
    
    if nuevo_estado in dict(SolicitudAdmision.ESTADOS_CHOICES):
        estado_anterior = solicitud.estado
        solicitud.estado = nuevo_estado
        solicitud.save()
        
        # Enviar notificación por email
        try:
            enviar_notificacion_cambio_estado(
                solicitud, 
                estado_anterior
            )
            messages.success(
                request, 
                f'Estado cambiado a "{solicitud.get_estado_display()}" y notificación enviada.'
            )
        except Exception as e:
            messages.warning(
                request, 
                f'Estado cambiado pero no se pudo enviar la notificación: {str(e)}'
            )
    else:
        messages.error(request, 'Estado no válido.')
    
    return redirect('admision:admin_ver_solicitud_publico', folio=folio)


@login_required
def admin_generar_ficha_publico(request, folio):
    if not _is_admin_user(request.user):
        messages.error(request, 'Acceso restringido: solo personal de Servicios Escolares.')
        return redirect('datos_academicos:servicios_login')
    """Generar ficha de admisión para una solicitud"""
    solicitud = get_object_or_404(SolicitudAdmision, folio=folio)
    
    if solicitud.estado != 'aceptada':
        messages.error(request, 'Solo se pueden generar fichas para solicitudes aceptadas.')
        return redirect('admision:admin_ver_solicitud_publico', folio=folio)
    
    try:
        # Verificar si ya existe una ficha
        if FichaAdmision.objects.filter(solicitud=solicitud).exists():
            messages.info(request, 'La ficha ya existe. Se reenvió por email.')
            enviar_ficha_por_email(solicitud)
        else:
            # Crear ficha y generar PDF
            ficha = FichaAdmision.objects.create(
                solicitud=solicitud,
                generada_por=request.user
            )
            
            from .utils import generar_ficha_admision_pdf
            from django.core.files.base import ContentFile
            pdf_content = generar_ficha_admision_pdf(solicitud)
            filename = f"ficha_{ficha.numero_ficha}.pdf"
            ficha.archivo_pdf.save(filename, ContentFile(pdf_content), save=True)
            
            # Enviar por email
            enviar_ficha_por_email(solicitud)
            messages.success(request, 'Ficha generada y enviada por email exitosamente.')
            
    except Exception as e:
        messages.error(request, f'Error al generar la ficha: {str(e)}')
    
    return redirect('admision:admin_ver_solicitud_publico', folio=folio)


@login_required
def admin_exportar_solicitudes(request):
    if not _is_admin_user(request.user):
        messages.error(request, 'Acceso restringido: solo personal de Servicios Escolares.')
        return redirect('datos_academicos:servicios_login')
    """Exportar solicitudes a Excel"""
    formato = request.GET.get('formato', 'excel')
    
    # Aplicar los mismos filtros que en la lista
    periodo_id = request.GET.get('periodo')
    estado = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    solicitudes = SolicitudAdmision.objects.all().order_by('-fecha_registro')
    
    if periodo_id:
        solicitudes = solicitudes.filter(periodo_id=periodo_id)
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_registro__date__gte=fecha_desde_dt)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_registro__date__lte=fecha_hasta_dt)
        except ValueError:
            pass
    
    if formato == 'csv':
        return _exportar_csv(solicitudes)
    else:
        return _exportar_excel(solicitudes)


def _exportar_csv(solicitudes):
    """Exportar a CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="solicitudes_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Folio', 'CURP', 'Email', 'Nombre Completo', 'Estado', 
        'Carrera de Interés', 'Fecha de Registro', 'Teléfono'
    ])
    
    for solicitud in solicitudes:
        writer.writerow([
            solicitud.folio,
            solicitud.curp,
            solicitud.email,
            solicitud.nombre_completo,
            solicitud.get_estado_display(),
            solicitud.obtener_respuesta('carrera_interes', ''),
            solicitud.fecha_registro.strftime('%Y-%m-%d %H:%M'),
            solicitud.obtener_respuesta('telefono', ''),
        ])
    
    return response


def _exportar_excel(solicitudes):
    """Exportar a Excel con formato"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitudes de Admisión"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'Folio', 'CURP', 'Email', 'Nombre', 'Apellido Paterno', 'Apellido Materno',
        'Estado', 'Carrera de Interés', 'Modalidad', 'Fecha de Nacimiento',
        'Género', 'Teléfono', 'Estado (Dirección)', 'Municipio',
        'Escuela de Procedencia', 'Promedio', 'Año de Egreso',
        'Fecha de Registro'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, solicitud in enumerate(solicitudes, 2):
        respuestas = solicitud.respuestas_json or {}
        
        data = [
            solicitud.folio,
            solicitud.curp,
            solicitud.email,
            respuestas.get('nombre', ''),
            respuestas.get('apellido_paterno', ''),
            respuestas.get('apellido_materno', ''),
            solicitud.get_estado_display(),
            respuestas.get('carrera_interes', ''),
            respuestas.get('modalidad', ''),
            respuestas.get('fecha_nacimiento', ''),
            respuestas.get('genero', ''),
            respuestas.get('telefono', ''),
            respuestas.get('estado', ''),
            respuestas.get('municipio', ''),
            respuestas.get('escuela_procedencia', ''),
            respuestas.get('promedio', ''),
            respuestas.get('ano_egreso', ''),
            solicitud.fecha_registro.strftime('%Y-%m-%d %H:%M'),
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="solicitudes_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    return response


@login_required
def admin_estadisticas_avanzadas(request):
    if not _is_admin_user(request.user):
        messages.error(request, 'Acceso restringido: solo personal de Servicios Escolares.')
        return redirect('datos_academicos:servicios_login')
    """Estadísticas avanzadas y gráficos"""
    periodo_actual = PeriodoAdmision.objects.filter(activo=True).first()
    
    if not periodo_actual:
        return render(request, 'admision/admin/estadisticas_avanzadas.html', {
            'error': 'No hay período de admisión activo'
        })
    
    solicitudes = SolicitudAdmision.objects.filter(periodo=periodo_actual)
    
    # Estadísticas por estado
    stats_estado = list(solicitudes.values('estado').annotate(
        count=Count('id')
    ).order_by('estado'))
    
    # Estadísticas por género
    stats_genero = []
    for genero in ['M', 'F', 'Otro']:
        count = solicitudes.extra(
            where=["JSON_EXTRACT(respuestas_json, '$.genero') = %s"],
            params=[genero]
        ).count()
        if count > 0:
            stats_genero.append({'genero': genero, 'count': count})
    
    # Estadísticas por rango de edad
    stats_edad = []
    rangos_edad = [
        ('17-18', 17, 18),
        ('19-20', 19, 20),
        ('21-22', 21, 22),
        ('23+', 23, 100)
    ]
    
    for rango, min_edad, max_edad in rangos_edad:
        # Calcular fechas límite
        fecha_max = timezone.now().date() - timedelta(days=min_edad * 365)
        fecha_min = timezone.now().date() - timedelta(days=(max_edad + 1) * 365)
        
        count = solicitudes.extra(
            where=["JSON_EXTRACT(respuestas_json, '$.fecha_nacimiento') BETWEEN %s AND %s"],
            params=[fecha_min.strftime('%Y-%m-%d'), fecha_max.strftime('%Y-%m-%d')]
        ).count()
        
        if count > 0:
            stats_edad.append({'rango': rango, 'count': count})
    
    # Registros por día (últimos 30 días)
    registros_por_dia = []
    for i in range(30):
        fecha = timezone.now().date() - timedelta(days=29-i)
        count = solicitudes.filter(fecha_registro__date=fecha).count()
        registros_por_dia.append({
            'fecha': fecha.strftime('%Y-%m-%d'),
            'count': count
        })
    
    context = {
        'periodo_actual': periodo_actual,
        'total_solicitudes': solicitudes.count(),
        'stats_estado': stats_estado,
        'stats_genero': stats_genero,
        'stats_edad': stats_edad,
        'registros_por_dia': registros_por_dia,
    }
    
    return render(request, 'admision/admin/estadisticas_avanzadas.html', context)


@csrf_exempt
def admin_accion_masiva(request):
    if not request.user.is_authenticated or not _is_admin_user(request.user):
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)
    """Realizar acciones masivas sobre solicitudes
    - Soporta selección por folios/emails o por filtros actuales
    - Restringe cambios de estado a 'aceptada' -> ('aceptada'|'rechazado')
    - Envía notificaciones por email al cambiar estado
    """
    # Autenticación y autorización explícitas para respuestas JSON
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'No autenticado'}, status=401)
    is_servicios = (
        request.user.is_staff
        or request.user.is_superuser
        or request.user.groups.filter(name__in=['ServiciosEscolares', 'Servicios Escolares']).exists()
    )
    if not is_servicios:
        return JsonResponse({'success': False, 'error': 'Acceso restringido: solo Servicios Escolares o staff'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    try:
        data = json.loads(request.body)
        accion = data.get('accion')
        apply_to = data.get('apply_to', 'selected')  # 'selected' | 'filtered'
        folios = data.get('folios', [])
        emails = data.get('emails', [])
        filtros_payload = data.get('filtros', {})

        queryset = SolicitudAdmision.objects.all()
        solicitudes = queryset

        if apply_to == 'selected':
            if not folios and not emails:
                return JsonResponse({'success': False, 'error': 'No se proporcionaron folios ni correos'})
            filtros_q = Q()
            if folios:
                filtros_q |= Q(folio__in=folios)
            if emails:
                filtros_q |= Q(email__in=emails)
            solicitudes = queryset.filter(filtros_q)
        else:  # apply_to == 'filtered'
            # Filtros similares a admin_solicitudes_publico
            periodo = filtros_payload.get('periodo') or filtros_payload.get('periodo_id')
            estado = filtros_payload.get('estado')
            busqueda = filtros_payload.get('busqueda') or filtros_payload.get('q')
            fecha_desde = filtros_payload.get('fecha_desde')
            fecha_hasta = filtros_payload.get('fecha_hasta')
            carrera = filtros_payload.get('carrera')

            if periodo:
                solicitudes = solicitudes.filter(periodo_id=periodo)
            if estado:
                solicitudes = solicitudes.filter(estado=estado)
            if busqueda:
                solicitudes = solicitudes.filter(
                    Q(folio__icontains=busqueda) |
                    Q(curp__icontains=busqueda) |
                    Q(email__icontains=busqueda) |
                    Q(respuestas_json__nombre__icontains=busqueda) |
                    Q(respuestas_json__apellido_paterno__icontains=busqueda) |
                    Q(respuestas_json__apellido_materno__icontains=busqueda)
                )
            if fecha_desde:
                try:
                    fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                    solicitudes = solicitudes.filter(fecha_registro__date__gte=fecha_desde_dt)
                except ValueError:
                    pass
            if fecha_hasta:
                try:
                    fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                    solicitudes = solicitudes.filter(fecha_registro__date__lte=fecha_hasta_dt)
                except ValueError:
                    pass
            if carrera:
                solicitudes = solicitudes.extra(
                    where=["JSON_EXTRACT(respuestas_json, '$.carrera_interes') = %s"],
                    params=[carrera]
                )

        if accion == 'cambiar_estado':
            nuevo_estado = data.get('nuevo_estado')
            if nuevo_estado not in dict(SolicitudAdmision.ESTADOS):
                return JsonResponse({'success': False, 'error': 'Estado no válido'})

            # Restricción: solo cambiar si el estado anterior fue 'aceptada'
            candidatos = solicitudes.filter(estado='aceptada')
            if not candidatos.exists():
                return JsonResponse({'success': False, 'error': 'No hay solicitudes en estado aceptada para actualizar'})

            actualizados = 0
            errores_envio = []
            from .email_utils import enviar_notificacion_cambio_estado
            for s in candidatos:
                anterior = s.estado
                s.estado = nuevo_estado
                s.save()
                actualizados += 1
                try:
                    ok = enviar_notificacion_cambio_estado(s, anterior)
                    if not ok:
                        errores_envio.append(f"Folio {s.folio}: envío fallido")
                except Exception as e:
                    errores_envio.append(str(e))

            nombre_estado = dict(SolicitudAdmision.ESTADOS).get(nuevo_estado, nuevo_estado)
            resp = {
                'success': True,
                'message': f'{actualizados} solicitudes actualizadas a "{nombre_estado}"',
                'updated_count': actualizados,
                'email_errors_count': len(errores_envio),
                'email_success_count': max(actualizados - len(errores_envio), 0)
            }
            if errores_envio:
                resp['warning'] = f'Errores de envío: {len(errores_envio)}'
            return JsonResponse(resp)

        elif accion == 'generar_fichas':
            fichas_generadas = 0
            for solicitud in solicitudes.filter(estado='aceptada'):
                ficha, created = FichaAdmision.objects.get_or_create(
                    solicitud=solicitud,
                    defaults={
                        'numero_ficha': FichaAdmision.generar_numero_ficha(),
                        'fecha_examen': timezone.now().date() + timedelta(days=30),
                        'hora_examen': timezone.now().time().replace(hour=9, minute=0),
                        'lugar_examen': 'Aula Magna - Edificio Principal',
                    }
                )
                if created:
                    fichas_generadas += 1
            return JsonResponse({'success': True, 'message': f'{fichas_generadas} fichas generadas'})

        return JsonResponse({'success': False, 'error': 'Acción no válida'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})